from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .models import ForecastRow, N4WorkloadRow, RefuelPlanRow, SourceStatus


STATUS_FILL = {
    "CRITICAL": "C00000",
    "WARNING": "FFC000",
    "OK": "70AD47",
    "NO_DATA": "A6A6A6",
}


def export_dashboard(
    output_path: Path,
    run_at: datetime,
    forecast_rows: list[ForecastRow],
    plan_rows: list[RefuelPlanRow],
    workload_rows: list[N4WorkloadRow],
    source_statuses: list[SourceStatus],
    warnings: list[str],
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    write_dashboard(ws, run_at, forecast_rows, plan_rows, warnings)
    write_forecast(wb.create_sheet("RTG_Forecast"), forecast_rows)
    write_refuel_plan(wb.create_sheet("Refuel_Plan"), plan_rows)
    write_workload(wb.create_sheet("N4_Workload"), workload_rows)
    write_source_status(wb.create_sheet("Source_Status"), source_statuses, warnings)
    wb.save(output_path)
    return output_path


def write_dashboard(ws: Worksheet, run_at: datetime, forecasts: list[ForecastRow], plan: list[RefuelPlanRow], warnings: list[str]) -> None:
    ws["A1"] = "FUEL ALERT RTG"
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:H1")
    ws["A2"] = "Run at"
    ws["B2"] = run_at
    ws["B2"].number_format = "yyyy-mm-dd hh:mm"
    counts = count_statuses(forecasts)
    for index, status in enumerate(["CRITICAL", "WARNING", "OK", "NO_DATA"], start=1):
        col = 1 + (index - 1) * 2
        ws.cell(row=4, column=col, value=status)
        ws.cell(row=5, column=col, value=counts.get(status, 0))
        ws.cell(row=4, column=col).font = Font(bold=True, color="FFFFFF" if status != "WARNING" else "000000")
        ws.cell(row=4, column=col).fill = PatternFill("solid", fgColor=STATUS_FILL[status])
        ws.cell(row=5, column=col).font = Font(size=16, bold=True)
    ws["A8"] = "Top RTG can do dau"
    ws["A8"].font = Font(size=13, bold=True)
    write_table(
        ws,
        9,
        ["Rank", "RTG", "Status", "Current %", "Liters to full", "Stop time", "Vessel", "ETB", "Reason"],
        [
            [
                row.rank,
                row.equipment,
                row.status,
                round_or_blank(row.current_pct),
                row.liters_to_full,
                row.time_to_stop,
                row.vessel_name,
                row.etb,
                row.reason,
            ]
            for row in plan[:10]
        ],
    )
    warn_start = 23
    ws.cell(row=warn_start, column=1, value="Canh bao du lieu").font = Font(size=13, bold=True)
    for offset, warning in enumerate(warnings or ["Khong co canh bao du lieu"], start=1):
        ws.cell(row=warn_start + offset, column=1, value=warning)
    finish_sheet(ws)


def write_forecast(ws: Worksheet, rows: list[ForecastRow]) -> None:
    write_table(
        ws,
        1,
        [
            "Equipment",
            "Last checklist",
            "Last level %",
            "Current liters",
            "Current %",
            "Time to 25%",
            "Time to 15%",
            "Hours to 25%",
            "Hours to 15%",
            "Status",
            "Checked by",
            "Note",
        ],
        [
            [
                row.equipment,
                row.last_check,
                round_or_blank(row.last_level_pct),
                round_or_blank(row.current_liters),
                round_or_blank(row.current_pct),
                row.time_to_warning,
                row.time_to_stop,
                round_or_blank(row.hours_to_warning),
                round_or_blank(row.hours_to_stop),
                row.status,
                row.checked_by,
                row.note,
            ]
            for row in rows
        ],
    )
    color_status_column(ws, 10, 2, len(rows) + 1)
    finish_sheet(ws)


def write_refuel_plan(ws: Worksheet, rows: list[RefuelPlanRow]) -> None:
    write_table(
        ws,
        1,
        ["Rank", "Equipment", "Status", "Current %", "Current liters", "Liters to full", "Time to 15%", "Visit", "Vessel", "ETB", "ETD", "N4 containers", "Reason"],
        [
            [
                row.rank,
                row.equipment,
                row.status,
                round_or_blank(row.current_pct),
                round_or_blank(row.current_liters),
                row.liters_to_full,
                row.time_to_stop,
                row.linked_visit,
                row.vessel_name,
                row.etb,
                row.etd,
                row.workload_containers,
                row.reason,
            ]
            for row in rows
        ],
    )
    color_status_column(ws, 3, 2, len(rows) + 1)
    finish_sheet(ws)


def write_workload(ws: Worksheet, rows: list[N4WorkloadRow]) -> None:
    write_table(
        ws,
        1,
        ["Visit", "Vessel", "Service", "Transit", "Current position", "Containers"],
        [[row.visit_code, row.vessel_name, row.service, row.transit, row.current_position, row.container_count] for row in rows],
    )
    finish_sheet(ws)


def write_source_status(ws: Worksheet, rows: list[SourceStatus], warnings: list[str]) -> None:
    write_table(
        ws,
        1,
        ["Source", "Path", "Exists", "Rows", "Last modified", "Status", "Message"],
        [[row.source, str(row.path or ""), row.exists, row.rows, row.last_modified, row.status, row.message] for row in rows],
    )
    start = len(rows) + 4
    ws.cell(row=start, column=1, value="Warnings").font = Font(bold=True)
    for index, warning in enumerate(warnings or ["No warnings"], start=start + 1):
        ws.cell(row=index, column=1, value=warning)
    finish_sheet(ws)


def write_table(ws: Worksheet, start_row: int, headers: list[str], rows: list[list[Any]]) -> None:
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=index, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_index, row in enumerate(rows, start=start_row + 1):
        for col_index, value in enumerate(row, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            if isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd hh:mm"
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=start_row, max_row=max(start_row + len(rows), start_row), min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")


def finish_sheet(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 10
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[letter].width = max_len + 2


def color_status_column(ws: Worksheet, column: int, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=column)
        fill = STATUS_FILL.get(str(cell.value), "")
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(bold=True, color="FFFFFF" if str(cell.value) != "WARNING" else "000000")


def count_statuses(rows: list[ForecastRow]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        counts[row.status] = counts.get(row.status, 0) + 1
    return counts


def round_or_blank(value: float | None, digits: int = 1) -> float | str:
    return "" if value is None else round(value, digits)
