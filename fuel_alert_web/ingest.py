from __future__ import annotations

import csv
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .models import FuelReading, N4WorkloadRow, SourceStatus
from .utils import clean_visit, normalize_header, normalize_rtg_name, parse_datetime, parse_float


def make_source_status(source: str, path: Path | None, rows: int, status: str, message: str) -> SourceStatus:
    if path is None:
        return SourceStatus(source=source, path=None, exists=False, rows=0, status=status, message=message)
    exists = path.exists()
    last_modified = datetime.fromtimestamp(path.stat().st_mtime) if exists else None
    return SourceStatus(source=source, path=path, exists=exists, rows=rows, last_modified=last_modified, status=status, message=message)


def read_fuel_readings(
    workbook_path: Path | None,
    sheet_name: str,
    equipment: list[str],
) -> tuple[dict[str, FuelReading], SourceStatus, list[str]]:
    if workbook_path is None:
        return {}, make_source_status("Fuel workbook", None, 0, "ERROR", "Chua upload Fuel level .xlsx"), [
            "Chua upload Fuel level .xlsx"
        ]
    if not workbook_path.exists():
        return {}, make_source_status("Fuel workbook", workbook_path, 0, "ERROR", "File khong ton tai"), [
            f"Khong tim thay file fuel workbook: {workbook_path}"
        ]

    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            return {}, make_source_status("Fuel workbook", workbook_path, 0, "ERROR", f"Khong co sheet {sheet_name}"), [
                f"Workbook khong co sheet {sheet_name}"
            ]
        ws = wb[sheet_name]
        raw_headers = [str(cell or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        headers = {normalize_header(header): index for index, header in enumerate(raw_headers)}
        aliases = {
            "completion_time": ["completion_time"],
            "equipment": ["ten_thiet_bi", "equipment", "ten_thiet_b"],
            "group": ["nhom_thiet_bi", "group", "nhom_thiet_b"],
            "fuel": ["so_dau", "fuel", "fuel_level"],
            "checked_by": ["cmit_name", "checked_by"],
        }
        indexes: dict[str, int] = {}
        for key, names in aliases.items():
            for name in names:
                if name in headers:
                    indexes[key] = headers[name]
                    break
        missing = [key for key in ["completion_time", "equipment", "group", "fuel"] if key not in indexes]
        if missing:
            return {}, make_source_status("Fuel workbook", workbook_path, max(ws.max_row - 1, 0), "ERROR", "Thieu cot"), [
                "Linked_data thieu cot: " + ", ".join(missing)
            ]

        allowed = {normalize_rtg_name(item) for item in equipment}
        readings: dict[str, FuelReading] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            group = str(row[indexes["group"]] or "").strip().upper()
            if group != "RTG":
                continue
            rtg = normalize_rtg_name(row[indexes["equipment"]])
            if rtg not in allowed:
                continue
            checked_at = parse_datetime(row[indexes["completion_time"]])
            level_pct = parse_float(row[indexes["fuel"]], None)
            if checked_at is None or level_pct is None:
                continue
            checked_by = ""
            if "checked_by" in indexes and row[indexes["checked_by"]] not in (None, ""):
                checked_by = str(row[indexes["checked_by"]]).strip()
            current = readings.get(rtg)
            if current is None or current.last_check is None or checked_at > current.last_check:
                readings[rtg] = FuelReading(rtg, checked_at, level_pct, checked_by)

        warnings = [f"Khong co checklist hop le cho {rtg}" for rtg in sorted(allowed) if rtg not in readings]
        return readings, make_source_status("Fuel workbook", workbook_path, max(ws.max_row - 1, 0), "OK", f"Doc {len(readings)} RTG"), warnings
    finally:
        wb.close()


def read_n4_workload(txt_path: Path | None, limit: int = 500) -> tuple[list[N4WorkloadRow], SourceStatus, list[str]]:
    if txt_path is None:
        return [], make_source_status("N4 TXT", None, 0, "WARNING", "Chua upload TXT N4"), ["Chua upload TXT N4"]
    if not txt_path.exists():
        return [], make_source_status("N4 TXT", txt_path, 0, "ERROR", "File khong ton tai"), [
            f"Khong tim thay file N4 TXT: {txt_path}"
        ]

    counts: Counter[tuple[str, str, str, str, str]] = Counter()
    total_rows = 0
    bad_rows = 0
    with txt_path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        for record in reader:
            total_rows += 1
            if not record or "Container No." not in record:
                bad_rows += 1
                continue
            visit = clean_visit(record.get("O/B Carrier") or record.get("Inbound Carrier") or "")
            vessel = str(record.get("O/B Carrier Name") or record.get("Inbound Carrier Name") or "").strip()
            service = str(record.get("Service") or "").strip()
            transit = str(record.get("Transit") or "").strip()
            position = str(record.get("Current position") or "").strip()
            counts[(visit, vessel, service, transit, position)] += 1

    rows = [
        N4WorkloadRow(key[0], key[1], key[2], key[3], key[4], count)
        for key, count in counts.most_common(limit)
    ]
    warnings = [f"N4 TXT co {bad_rows} dong khong dung dinh dang TSV/container"] if bad_rows else []
    return rows, make_source_status("N4 TXT", txt_path, total_rows, "OK", f"Doc {total_rows} dong, tong hop {len(rows)} nhom"), warnings


def workload_by_visit(rows: list[N4WorkloadRow]) -> dict[str, int]:
    totals: defaultdict[str, int] = defaultdict(int)
    for row in rows:
        totals[row.visit_code] += row.container_count
    return dict(totals)
